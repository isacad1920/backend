"""
Financial report export service.
"""
import logging
import io
import os
import json
import csv
from datetime import datetime, date
from typing import Optional, Dict, Any, Union
from decimal import Decimal
from generated.prisma import Prisma

logger = logging.getLogger(__name__)

class ExportService:
    """Service for exporting financial reports."""
    
    def __init__(self, db: Prisma):
        """Initialize export service.
        
        Args:
            db: Prisma database client
        """
        self.db = db
    
    async def export_financial_report(
        self,
        report_type: str,
        format: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        branch_id: Optional[int] = None,
        current_user: Dict[str, Any] = None
    ) -> Dict[str, Any]:
        """Export financial report in specified format.
        
        Args:
            report_type: Type of report ('income-statement', 'balance-sheet', 'cash-flow', 'tax-report')
            format: Export format ('json', 'csv', 'pdf', 'excel')
            start_date: Period start date
            end_date: Period end date
            branch_id: Optional branch filter
            current_user: Current authenticated user
            
        Returns:
            Export result with file path/data and metadata
        """
        try:
            # Check permissions
            if not self._check_export_permission(current_user):
                raise ValueError("Insufficient permissions to export financial reports")
            
            # Import report services when needed
            from .report_service import ReportService
            report_service = ReportService(self.db)
            
            # Set default dates if not provided
            if not start_date:
                start_date = date.today().replace(day=1)
            if not end_date:
                end_date = date.today()
            
            # Generate report data based on type
            if report_type == 'income-statement':
                report_data = await report_service.generate_income_statement(
                    start_date, end_date, branch_id, current_user
                )
            elif report_type == 'balance-sheet':
                report_data = await report_service.generate_balance_sheet(
                    end_date, branch_id, current_user
                )
            elif report_type == 'cash-flow':
                report_data = await report_service.generate_cash_flow_statement(
                    start_date, end_date, branch_id, current_user
                )
            elif report_type == 'tax-report':
                report_data = await report_service.generate_tax_report(
                    start_date, end_date, branch_id, current_user
                )
            else:
                raise ValueError(f"Unsupported report type: {report_type}")
            
            # Export in specified format
            if format.lower() == 'json':
                return self._export_as_json(report_data, report_type)
            elif format.lower() == 'csv':
                return await self._export_as_csv(report_data, report_type)
            elif format.lower() == 'pdf':
                return await self._export_as_pdf(report_data, report_type)
            elif format.lower() == 'excel':
                return await self._export_as_excel(report_data, report_type)
            else:
                raise ValueError(f"Unsupported export format: {format}")
            
        except Exception as e:
            logger.error(f"Error exporting financial report: {e}")
            raise
    
    def _export_as_json(self, report_data: Any, report_type: str) -> Dict[str, Any]:
        """Export report data as JSON.
        
        Args:
            report_data: Report data to export
            report_type: Type of report being exported
            
        Returns:
            JSON export result
        """
        try:
            # Convert Pydantic model to dict
            if hasattr(report_data, 'dict'):
                data_dict = report_data.dict()
            else:
                data_dict = report_data
            
            # Handle Decimal serialization
            def decimal_serializer(obj):
                if isinstance(obj, Decimal):
                    return float(obj)
                elif isinstance(obj, (date, datetime)):
                    return obj.isoformat()
                raise TypeError(f"Object of type {type(obj)} is not JSON serializable")
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{report_type}_{timestamp}.json"
            filepath = os.path.join("exports", filename)
            
            # Ensure exports directory exists
            os.makedirs("exports", exist_ok=True)
            
            # Write JSON file
            with open(filepath, 'w') as f:
                json.dump(data_dict, f, indent=2, default=decimal_serializer)
            
            return {
                'success': True,
                'format': 'json',
                'filename': filename,
                'filepath': filepath,
                'size': os.path.getsize(filepath),
                'mime_type': 'application/json',
                'generated_at': datetime.utcnow().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Error exporting as JSON: {e}")
            raise
    
    async def _export_as_csv(self, report_data: Any, report_type: str) -> Dict[str, Any]:
        """Export report data as CSV.
        
        Args:
            report_data: Report data to export
            report_type: Type of report being exported
            
        Returns:
            CSV export result
        """
        try:
            # Convert to dict if needed
            if hasattr(report_data, 'dict'):
                data_dict = report_data.dict()
            else:
                data_dict = report_data
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{report_type}_{timestamp}.csv"
            filepath = os.path.join("exports", filename)
            
            # Ensure exports directory exists
            os.makedirs("exports", exist_ok=True)
            
            # Write CSV file based on report type
            with open(filepath, 'w', newline='') as f:
                writer = csv.writer(f)
                
                if report_type == 'income-statement':
                    self._write_income_statement_csv(writer, data_dict)
                elif report_type == 'balance-sheet':
                    self._write_balance_sheet_csv(writer, data_dict)
                elif report_type == 'cash-flow':
                    self._write_cash_flow_csv(writer, data_dict)
                elif report_type == 'tax-report':
                    self._write_tax_report_csv(writer, data_dict)
                else:
                    # Generic CSV export
                    self._write_generic_csv(writer, data_dict)
            
            return {
                'success': True,
                'format': 'csv',
                'filename': filename,
                'filepath': filepath,
                'size': os.path.getsize(filepath),
                'mime_type': 'text/csv',
                'generated_at': datetime.utcnow().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Error exporting as CSV: {e}")
            raise
    
    async def _export_as_pdf(self, report_data: Any, report_type: str) -> Dict[str, Any]:
        """Export report data as PDF.
        
        Args:
            report_data: Report data to export
            report_type: Type of report being exported
            
        Returns:
            PDF export result
        """
        try:
            # Import reportlab when needed
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.units import inch
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{report_type}_{timestamp}.pdf"
            filepath = os.path.join("exports", filename)
            
            # Ensure exports directory exists
            os.makedirs("exports", exist_ok=True)
            
            # Create PDF document
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # Add title
            title = Paragraph(f"{report_type.replace('-', ' ').title()} Report", styles['Title'])
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Convert to dict if needed
            if hasattr(report_data, 'dict'):
                data_dict = report_data.dict()
            else:
                data_dict = report_data
            
            # Add report content based on type
            if report_type == 'income-statement':
                self._add_income_statement_pdf_content(story, data_dict, styles)
            elif report_type == 'balance-sheet':
                self._add_balance_sheet_pdf_content(story, data_dict, styles)
            elif report_type == 'cash-flow':
                self._add_cash_flow_pdf_content(story, data_dict, styles)
            elif report_type == 'tax-report':
                self._add_tax_report_pdf_content(story, data_dict, styles)
            else:
                # Generic PDF content
                self._add_generic_pdf_content(story, data_dict, styles)
            
            # Build PDF
            doc.build(story)
            
            return {
                'success': True,
                'format': 'pdf',
                'filename': filename,
                'filepath': filepath,
                'size': os.path.getsize(filepath),
                'mime_type': 'application/pdf',
                'generated_at': datetime.utcnow().isoformat()
            }
            
        except ImportError:
            raise ValueError("PDF export requires reportlab package")
        except Exception as e:
            logger.error(f"Error exporting as PDF: {e}")
            raise
    
    async def _export_as_excel(self, report_data: Any, report_type: str) -> Dict[str, Any]:
        """Export report data as Excel file.
        
        Args:
            report_data: Report data to export
            report_type: Type of report being exported
            
        Returns:
            Excel export result
        """
        try:
            # Import openpyxl when needed
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{report_type}_{timestamp}.xlsx"
            filepath = os.path.join("exports", filename)
            
            # Ensure exports directory exists
            os.makedirs("exports", exist_ok=True)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = report_type.replace('-', ' ').title()
            
            # Convert to dict if needed
            if hasattr(report_data, 'dict'):
                data_dict = report_data.dict()
            else:
                data_dict = report_data
            
            # Add content based on report type
            if report_type == 'income-statement':
                self._add_income_statement_excel_content(ws, data_dict)
            elif report_type == 'balance-sheet':
                self._add_balance_sheet_excel_content(ws, data_dict)
            elif report_type == 'cash-flow':
                self._add_cash_flow_excel_content(ws, data_dict)
            elif report_type == 'tax-report':
                self._add_tax_report_excel_content(ws, data_dict)
            else:
                # Generic Excel content
                self._add_generic_excel_content(ws, data_dict)
            
            # Save workbook
            wb.save(filepath)
            
            return {
                'success': True,
                'format': 'excel',
                'filename': filename,
                'filepath': filepath,
                'size': os.path.getsize(filepath),
                'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'generated_at': datetime.utcnow().isoformat()
            }
            
        except ImportError:
            raise ValueError("Excel export requires openpyxl package")
        except Exception as e:
            logger.error(f"Error exporting as Excel: {e}")
            raise
    
    # CSV writing helper methods
    def _write_income_statement_csv(self, writer, data_dict):
        """Write income statement data to CSV."""
        writer.writerow(['Income Statement'])
        writer.writerow([f"Period: {data_dict.get('period_start')} to {data_dict.get('period_end')}"])
        writer.writerow([])
        
        # Revenue section
        writer.writerow(['REVENUE'])
        for item in data_dict.get('revenue', []):
            writer.writerow([item.get('description'), item.get('amount')])
        writer.writerow(['Total Revenue', data_dict.get('total_revenue')])
        writer.writerow([])
        
        # COGS section
        writer.writerow(['COST OF GOODS SOLD'])
        for item in data_dict.get('cost_of_goods_sold', []):
            writer.writerow([item.get('description'), item.get('amount')])
        writer.writerow(['Total COGS', data_dict.get('total_cogs')])
        writer.writerow([])
        
        # Gross Profit
        writer.writerow(['Gross Profit', data_dict.get('gross_profit')])
        writer.writerow([])
        
        # Expenses
        writer.writerow(['EXPENSES'])
        for item in data_dict.get('expenses', []):
            writer.writerow([item.get('description'), item.get('amount')])
        writer.writerow(['Total Expenses', data_dict.get('total_expenses')])
        writer.writerow([])
        
        # Net Profit
        writer.writerow(['Net Profit', data_dict.get('net_profit')])
    
    def _write_balance_sheet_csv(self, writer, data_dict):
        """Write balance sheet data to CSV."""
        writer.writerow(['Balance Sheet'])
        writer.writerow([f"As of: {data_dict.get('as_of_date')}"])
        writer.writerow([])
        
        # Assets
        writer.writerow(['ASSETS'])
        for item in data_dict.get('assets', []):
            writer.writerow([item.get('description'), item.get('amount')])
        writer.writerow(['Total Assets', data_dict.get('total_assets')])
        writer.writerow([])
        
        # Liabilities
        writer.writerow(['LIABILITIES'])
        for item in data_dict.get('liabilities', []):
            writer.writerow([item.get('description'), item.get('amount')])
        writer.writerow(['Total Liabilities', data_dict.get('total_liabilities')])
        writer.writerow([])
        
        # Equity
        writer.writerow(['EQUITY'])
        for item in data_dict.get('equity', []):
            writer.writerow([item.get('description'), item.get('amount')])
        writer.writerow(['Total Equity', data_dict.get('total_equity')])
    
    def _write_cash_flow_csv(self, writer, data_dict):
        """Write cash flow data to CSV."""
        writer.writerow(['Cash Flow Statement'])
        writer.writerow([f"Period: {data_dict.get('period_start')} to {data_dict.get('period_end')}"])
        writer.writerow([])
        
        # Operating Activities
        writer.writerow(['OPERATING ACTIVITIES'])
        for item in data_dict.get('operating_activities', []):
            writer.writerow([item.get('description'), item.get('amount')])
        writer.writerow(['Net Cash from Operating', data_dict.get('operating_total')])
        writer.writerow([])
        
        # Investing Activities
        writer.writerow(['INVESTING ACTIVITIES'])
        for item in data_dict.get('investing_activities', []):
            writer.writerow([item.get('description'), item.get('amount')])
        writer.writerow(['Net Cash from Investing', data_dict.get('investing_total')])
        writer.writerow([])
        
        # Financing Activities
        writer.writerow(['FINANCING ACTIVITIES'])
        for item in data_dict.get('financing_activities', []):
            writer.writerow([item.get('description'), item.get('amount')])
        writer.writerow(['Net Cash from Financing', data_dict.get('financing_total')])
        writer.writerow([])
        
        # Net Change
        writer.writerow(['Net Change in Cash', data_dict.get('net_cash_flow')])
        writer.writerow(['Opening Balance', data_dict.get('opening_cash_balance')])
        writer.writerow(['Closing Balance', data_dict.get('closing_cash_balance')])
    
    def _write_tax_report_csv(self, writer, data_dict):
        """Write tax report data to CSV."""
        writer.writerow(['Tax Report'])
        writer.writerow([f"Period: {data_dict.get('period_start')} to {data_dict.get('period_end')}"])
        writer.writerow([])
        
        writer.writerow(['Taxable Income', data_dict.get('total_taxable_income')])
        writer.writerow(['Deductions', data_dict.get('total_deductions')])
        writer.writerow(['Adjusted Taxable Income', data_dict.get('adjusted_taxable_income')])
        writer.writerow(['Tax Rate', f"{data_dict.get('tax_rate', 0)*100}%"])
        writer.writerow(['Tax Liability', data_dict.get('tax_liability')])
        writer.writerow(['Payments Made', data_dict.get('total_payments')])
        writer.writerow(['Balance Due', data_dict.get('balance_due')])
    
    def _write_generic_csv(self, writer, data_dict):
        """Write generic data to CSV."""
        writer.writerow(['Financial Report'])
        writer.writerow([])
        
        for key, value in data_dict.items():
            if isinstance(value, (int, float, str)):
                writer.writerow([key, value])
    
    # PDF content helper methods
    def _add_income_statement_pdf_content(self, story, data_dict, styles):
        """Add income statement content to PDF."""
        from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        
        # Add period info
        period_text = f"Period: {data_dict.get('period_start')} to {data_dict.get('period_end')}"
        story.append(Paragraph(period_text, styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Create table data
        table_data = [
            ['Description', 'Amount'],
            ['REVENUE', ''],
        ]
        
        # Add revenue items
        for item in data_dict.get('revenue', []):
            table_data.append([item.get('description'), f"${item.get('amount', 0):,.2f}"])
        table_data.append(['Total Revenue', f"${data_dict.get('total_revenue', 0):,.2f}"])
        
        # Add blank row
        table_data.append(['', ''])
        
        # Add COGS
        table_data.append(['COST OF GOODS SOLD', ''])
        for item in data_dict.get('cost_of_goods_sold', []):
            table_data.append([item.get('description'), f"${item.get('amount', 0):,.2f}"])
        table_data.append(['Total COGS', f"${data_dict.get('total_cogs', 0):,.2f}"])
        
        # Add gross profit
        table_data.append(['', ''])
        table_data.append(['Gross Profit', f"${data_dict.get('gross_profit', 0):,.2f}"])
        
        # Create and style table
        table = Table(table_data, colWidths=[4*72, 2*72])  # 4 inches, 2 inches
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
    
    def _add_balance_sheet_pdf_content(self, story, data_dict, styles):
        """Add balance sheet content to PDF."""
        # Simplified implementation
        from reportlab.platypus import Paragraph
        story.append(Paragraph("Balance Sheet Content", styles['Normal']))
    
    def _add_cash_flow_pdf_content(self, story, data_dict, styles):
        """Add cash flow content to PDF."""
        # Simplified implementation
        from reportlab.platypus import Paragraph
        story.append(Paragraph("Cash Flow Content", styles['Normal']))
    
    def _add_tax_report_pdf_content(self, story, data_dict, styles):
        """Add tax report content to PDF."""
        # Simplified implementation
        from reportlab.platypus import Paragraph
        story.append(Paragraph("Tax Report Content", styles['Normal']))
    
    def _add_generic_pdf_content(self, story, data_dict, styles):
        """Add generic content to PDF."""
        from reportlab.platypus import Paragraph
        story.append(Paragraph("Generic Report Content", styles['Normal']))
    
    # Excel content helper methods
    def _add_income_statement_excel_content(self, ws, data_dict):
        """Add income statement content to Excel."""
        # Simplified implementation
        ws['A1'] = 'Income Statement'
        ws['A2'] = f"Period: {data_dict.get('period_start')} to {data_dict.get('period_end')}"
        
        row = 4
        ws[f'A{row}'] = 'Total Revenue'
        ws[f'B{row}'] = data_dict.get('total_revenue', 0)
        
        row += 1
        ws[f'A{row}'] = 'Total COGS'
        ws[f'B{row}'] = data_dict.get('total_cogs', 0)
        
        row += 1
        ws[f'A{row}'] = 'Gross Profit'
        ws[f'B{row}'] = data_dict.get('gross_profit', 0)
        
        row += 1
        ws[f'A{row}'] = 'Net Profit'
        ws[f'B{row}'] = data_dict.get('net_profit', 0)
    
    def _add_balance_sheet_excel_content(self, ws, data_dict):
        """Add balance sheet content to Excel."""
        ws['A1'] = 'Balance Sheet'
        ws['A2'] = f"As of: {data_dict.get('as_of_date')}"
    
    def _add_cash_flow_excel_content(self, ws, data_dict):
        """Add cash flow content to Excel."""
        ws['A1'] = 'Cash Flow Statement'
        ws['A2'] = f"Period: {data_dict.get('period_start')} to {data_dict.get('period_end')}"
    
    def _add_tax_report_excel_content(self, ws, data_dict):
        """Add tax report content to Excel."""
        ws['A1'] = 'Tax Report'
        ws['A2'] = f"Period: {data_dict.get('period_start')} to {data_dict.get('period_end')}"
    
    def _add_generic_excel_content(self, ws, data_dict):
        """Add generic content to Excel."""
        ws['A1'] = 'Financial Report'
        
        row = 3
        for key, value in data_dict.items():
            if isinstance(value, (int, float, str)):
                ws[f'A{row}'] = str(key)
                ws[f'B{row}'] = value
                row += 1
    
    def _check_export_permission(self, user: Dict[str, Any]) -> bool:
        """Check if user has permission to export financial reports.
        
        Args:
            user: Current user data
            
        Returns:
            True if permission granted, False otherwise
        """
        if not user:
            return False
        
        # Check if user has required permissions
        user_role = user.get('role', 'CASHIER')
        return user_role in ['MANAGER', 'ADMIN', 'ACCOUNTANT']
